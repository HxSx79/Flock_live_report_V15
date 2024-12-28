import openpyxl
from datetime import datetime
from typing import Dict
import os

class ExcelLogger:
    def __init__(self, filename="production_log.xlsx"):
        self.filename = filename
        self._ensure_file_exists()
        
    def _ensure_file_exists(self):
        """Create the Excel file with headers if it doesn't exist"""
        if not os.path.exists(self.filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = [
                "Timestamp", 
                "Line", 
                "Class Name", 
                "Program", 
                "Part Number", 
                "Description"
            ]
            ws.append(headers)
            wb.save(self.filename)
            print(f"Created new production log file: {self.filename}")
    
    def log_crossing(self, line_number: int, class_name: str, part_info: Dict):
        """Log a line crossing event to Excel"""
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            
            # Prepare row data
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row_data = [
                timestamp,
                f"Line {line_number}",
                class_name,
                part_info['program'],
                part_info['part_number'],
                part_info['description']
            ]
            
            # Insert new row after header
            ws.insert_rows(2)
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=2, column=col, value=value)
            
            wb.save(self.filename)
            print(f"Logged crossing event: Line {line_number}, {class_name}")
            
        except Exception as e:
            print(f"Error logging crossing event: {e}")
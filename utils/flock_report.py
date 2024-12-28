import openpyxl
from datetime import datetime
from typing import Dict
import os
from .bom_reader import BOMReader

class FlockReport:
    def __init__(self, filename="flock_report.xlsx"):
        self.filename = filename
        self.bom_reader = BOMReader()
        self._ensure_file_exists()
        
    def _ensure_file_exists(self):
        """Create the Excel file with headers if it doesn't exist"""
        if not os.path.exists(self.filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["Class Name", "Program", "Part Number", "Part Description", "Day", "Time"]
            ws.append(headers)
            wb.save(self.filename)
    
    def record_crossing(self, class_name: str):
        """Record a line crossing event in the Excel file"""
        # Get part information from BOM
        part_info = self.bom_reader.get_part_info(class_name)
        
        # Get current date and time
        now = datetime.now()
        day = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")
        
        # Prepare row data
        row_data = [
            class_name,
            part_info['program'],
            part_info['part_number'],
            part_info['description'],
            day,
            time
        ]
        
        # Open workbook and insert new row after header
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        ws.insert_rows(2)  # Insert after header row
        
        # Write data to the new row
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=2, column=col, value=value)
        
        # Save the workbook
        wb.save(self.filename)